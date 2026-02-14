# Data driven testing -> Data driven testing is testing of a functionality with different data set. Application are developed first and testing begins later
# Test driven testing -> Test driven testing uses testing first and development later.

import openpyxl
filename= "yourexcelpath"
workbk = openpyxl.load_workbook(filename=filename)
sheet = workbk['user_data']
rows = sheet.max_row # 5
cols = sheet.max_column # 2
for each_row in range(1,rows+1): # 1,6
    for each_col in range(1,cols+1): # 1,3
       print(sheet.cell(row=each_row,column=each_col).value) #1,1 # 1,2 # 2,1 # 2,2-----#5,1 and 5,2
       sheet.cell(row=each_row,column=3).value = 'read'
workbk.save(filename=filename)